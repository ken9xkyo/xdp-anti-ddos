#!/usr/bin/env python3
"""
XDP Anti-DDoS - Bộ kiểm thử toàn diện (Comprehensive Test Suite)

Sử dụng bpftool prog run + Scapy để kiểm tra tất cả nhánh xử lý
của chương trình XDP anti-DDoS.

Yêu cầu:
  - XDP program đã được load (make IFACE=<if> attach)
  - Pinned tại /sys/fs/bpf/xdp_anti_ddos/xdp_anti_ddos
  - scapy, openpyxl

Chạy: sudo python3 test_xdp_ddos.py
"""
import subprocess
import json
import os
import struct
import socket
from datetime import datetime
from scapy.all import Ether, IP, IPv6, UDP, TCP, ICMP, Dot1Q, raw

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[-] Thư viện openpyxl chưa được cài đặt. Hãy chạy: pip install openpyxl")
    exit(1)

# --- Các hằng số XDP Action ---
XDP_ABORTED = 0
XDP_DROP = 1
XDP_PASS = 2
XDP_TX = 3
XDP_REDIRECT = 4

# Map code ra string
ACTION_MAP = {
    XDP_ABORTED: "XDP_ABORTED",
    XDP_DROP: "XDP_DROP",
    XDP_PASS: "XDP_PASS",
    XDP_TX: "XDP_TX",
    XDP_REDIRECT: "XDP_REDIRECT"
}

PROG_PATH = "/sys/fs/bpf/xdp_anti_ddos/xdp_anti_ddos"

# Danh sách lưu trữ kết quả để xuất Excel
test_results = []

# =======================================================================
# HÀM HELPER QUẢN LÝ BPF MAP
# =======================================================================

def _get_map_id(map_name):
    """Tìm map ID từ tên map trong XDP program đang active."""
    try:
        result = subprocess.run(
            ['bpftool', 'prog', 'show', '-j'],
            capture_output=True, text=True, timeout=10
        )
        if result.returncode != 0:
            return None
        progs = json.loads(result.stdout)
        active_map_ids = []
        for prog in progs:
            if prog.get('type') == 'xdp' and 'xdp_anti_ddos' in prog.get('name', ''):
                active_map_ids = prog.get('map_ids', [])
                break

        result = subprocess.run(
            ['bpftool', 'map', 'show', '-j'],
            capture_output=True, text=True, timeout=10
        )
        if result.returncode != 0:
            return None
        maps = json.loads(result.stdout)
        name_prefix = map_name[:15] if len(map_name) > 15 else map_name
        if active_map_ids:
            for m in maps:
                mn = m.get('name', '')
                if (mn == map_name or mn.startswith(name_prefix)) and m.get('id') in active_map_ids:
                    return m.get('id')
        matching = [m for m in maps if m.get('name', '').startswith(name_prefix)]
        if matching:
            matching.sort(key=lambda x: x.get('id', 0), reverse=True)
            return matching[0].get('id')
    except Exception:
        pass
    return None


def _ip_to_hex_bytes(ip_str):
    """Chuyển IP string thành danh sách hex bytes."""
    packed = socket.inet_aton(ip_str)
    return [f'0x{b:02x}' for b in packed]


def _port_to_hex_bytes(port):
    """Chuyển port number thành hex bytes (little-endian u16)."""
    return [f'0x{port & 0xff:02x}', f'0x{(port >> 8) & 0xff:02x}']


def map_add_ip(map_name, ip_str):
    """Thêm 1 IP vào BPF map (whitelist hoặc blacklist)."""
    map_id = _get_map_id(map_name)
    if not map_id:
        print(f"    [WARN] Map '{map_name}' không tìm thấy, bỏ qua setup")
        return False
    hex_key = _ip_to_hex_bytes(ip_str)
    cmd = ['sudo', 'bpftool', 'map', 'update', 'id', str(map_id),
           'key'] + hex_key + ['value', '0x01']
    r = subprocess.run(cmd, capture_output=True, text=True)
    return r.returncode == 0


def map_delete_ip(map_name, ip_str):
    """Xoá 1 IP khỏi BPF map."""
    map_id = _get_map_id(map_name)
    if not map_id:
        return False
    hex_key = _ip_to_hex_bytes(ip_str)
    cmd = ['sudo', 'bpftool', 'map', 'delete', 'id', str(map_id),
           'key'] + hex_key
    r = subprocess.run(cmd, capture_output=True, text=True)
    return r.returncode == 0


def map_add_port(map_name, port):
    """Thêm 1 port vào BPF map (amp_ports_map)."""
    map_id = _get_map_id(map_name)
    if not map_id:
        print(f"    [WARN] Map '{map_name}' không tìm thấy, bỏ qua setup")
        return False
    hex_key = _port_to_hex_bytes(port)
    cmd = ['sudo', 'bpftool', 'map', 'update', 'id', str(map_id),
           'key'] + hex_key + ['value', '0x01']
    r = subprocess.run(cmd, capture_output=True, text=True)
    return r.returncode == 0


def map_delete_port(map_name, port):
    """Xoá 1 port khỏi BPF map."""
    map_id = _get_map_id(map_name)
    if not map_id:
        return False
    hex_key = _port_to_hex_bytes(port)
    cmd = ['sudo', 'bpftool', 'map', 'delete', 'id', str(map_id),
           'key'] + hex_key
    r = subprocess.run(cmd, capture_output=True, text=True)
    return r.returncode == 0


# =======================================================================
# HÀM CHẠY TEST
# =======================================================================

def run_xdp_test(tc_id, test_name, action_desc, expect_desc, pkt_bytes, expected_action):
    """
    Thực thi test, in log ra console và lưu dữ liệu vào test_results.
    """
    in_file = "tmp_pkt_in.bin"
    out_file = "tmp_pkt_out.bin"
    
    print(f"\n[▶] {tc_id}: {test_name}")
    
    # Chuẩn bị bản ghi dữ liệu
    record = {
        "TC_ID": tc_id,
        "Tên Test Case": test_name,
        "Hành Động": action_desc,
        "Kỳ Vọng": expect_desc,
        "Kích Thước Gói (Bytes)": len(pkt_bytes),
        "Mã Kỳ Vọng": ACTION_MAP.get(expected_action, str(expected_action)),
        "Mã Thực Tế": "N/A",
        "Thời Gian (ns)": 0,
        "Trạng Thái": "ERROR",
        "Ghi Chú": ""
    }
    
    with open(in_file, "wb") as f:
        f.write(pkt_bytes)
        
    cmd = [
        "sudo", "bpftool", "prog", "run",
        "pinned", PROG_PATH,
        "data_in", in_file,
        "data_out", out_file,
        "-j" 
    ]
    
    result = subprocess.run(cmd, capture_output=True, text=True)
    
    if result.returncode != 0:
        print(f"    └─ [✘] LỖI BPFTOOL: {result.stderr.strip()}")
        record["Ghi Chú"] = result.stderr.strip()
        test_results.append(record)
        return False
        
    try:
        output = json.loads(result.stdout)
        retval = output.get("retval")
        duration = output.get("duration", 0)
        
        record["Mã Thực Tế"] = ACTION_MAP.get(retval, str(retval))
        record["Thời Gian (ns)"] = duration
        
        if retval == expected_action:
            print(f"    └─ [✔] PASS ({duration} ns)")
            record["Trạng Thái"] = "PASS"
        else:
            print(f"    └─ [✘] FAIL (Thực tế: {record['Mã Thực Tế']})")
            record["Trạng Thái"] = "FAIL"
            record["Ghi Chú"] = "Sai lệch Action"
            
    except json.JSONDecodeError:
        print("    └─ [✘] Lỗi parse JSON")
        record["Ghi Chú"] = "Lỗi JSON"
    finally:
        test_results.append(record)
        if os.path.exists(in_file): os.remove(in_file)
        if os.path.exists(out_file): os.remove(out_file)

# =======================================================================
# HÀM XUẤT BÁO CÁO EXCEL CHUYÊN NGHIỆP
# =======================================================================

def export_excel_report(filename="XDP_Anti_DDoS_Test_Report.xlsx"):
    if not test_results:
        print("\n[!] Không có dữ liệu để xuất Excel.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Results"

    # Định nghĩa các Style
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    pass_fill = PatternFill(start_color="C4D79B", end_color="C4D79B", fill_type="solid")
    fail_fill = PatternFill(start_color="DA9694", end_color="DA9694", fill_type="solid")

    # Viết Header
    headers = list(test_results[0].keys())
    ws.append(headers)
    
    for col_num, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border

    # Ghi dữ liệu và format
    for row_idx, data in enumerate(test_results, 2):
        row_data = list(data.values())
        ws.append(row_data)
        
        for col_idx, cell in enumerate(ws[row_idx], 1):
            cell.border = thin_border
            if headers[col_idx-1] in ["TC_ID", "Kích Thước Gói (Bytes)", "Mã Kỳ Vọng", "Mã Thực Tế", "Thời Gian (ns)", "Trạng Thái"]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

            # Tô màu dựa trên trạng thái PASS/FAIL
            if headers[col_idx-1] == "Trạng Thái":
                cell.font = Font(bold=True)
                if cell.value == "PASS":
                    cell.fill = pass_fill
                    cell.font = Font(color="006100", bold=True)
                else:
                    cell.fill = fail_fill
                    cell.font = Font(color="9C0006", bold=True)

    # Tự động căn chỉnh độ rộng cột (Auto-fit)
    column_widths = {
        'A': 10,  # TC_ID
        'B': 40,  # Tên
        'C': 50,  # Hành Động
        'D': 50,  # Kỳ Vọng
        'E': 20,  # Kích thước
        'F': 15,  # Mã Kỳ Vọng
        'G': 15,  # Mã Thực Tế
        'H': 15,  # Thời gian
        'I': 15,  # Trạng Thái
        'J': 30   # Ghi Chú
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Đóng băng dòng tiêu đề để dễ cuộn chuột
    ws.freeze_panes = "A2"

    try:
        wb.save(filename)
        print(f"\n[INFO] Đã xuất báo cáo Excel thành công tại: {os.path.abspath(filename)}")
    except Exception as e:
        print(f"\n[ERROR] Không thể lưu file Excel: {e}")

# =======================================================================
# KỊCH BẢN KIỂM THỬ
# =======================================================================

def run_all_tests():
    print("🚀 BẮT ĐẦU CHẠY XDP ANTI-DDOS UNIT TEST SUITE...\n")

    # ===================================================================
    # NHÓM 1: PHÂN TÍCH GÓI TIN (Parsing Tests)
    # ===================================================================
    print("=" * 60)
    print("📦 NHÓM 1: PHÂN TÍCH GÓI TIN (Parsing)")
    print("=" * 60)

    run_xdp_test(
        "TC_1.1", "Gói tin quá nhỏ",
        "Bơm mảng 10 byte rác.",
        "Bắt lỗi (eth+1 > data_end) -> DROP",
        b"\x00" * 10, XDP_DROP
    )
    
    pkt_ipv6 = Ether() / IPv6() / UDP(sport=1234, dport=80)
    run_xdp_test(
        "TC_1.2", "Bỏ qua IPv6",
        "Craft gói IPv6 UDP.",
        "Không phải IPv4 -> PASS",
        raw(pkt_ipv6), XDP_PASS
    )
    
    pkt_trunc_ip = raw(Ether() / IP())[:24] 
    run_xdp_test(
        "TC_1.3", "IP Header bị cắt cụt",
        "Gửi gói IP thiếu byte (chỉ 24 byte, cần >= 34).",
        "Bắt lỗi kích thước IP header -> DROP",
        pkt_trunc_ip, XDP_DROP
    )
    
    pkt_vlan = Ether() / Dot1Q(vlan=100) / IP() / UDP(sport=10000, dport=80)
    run_xdp_test(
        "TC_1.4", "Bóc tách VLAN đơn (802.1Q)",
        "Gửi gói bọc VLAN tag 100.",
        "Bóc tag thành công, gói UDP hợp lệ -> PASS",
        raw(pkt_vlan), XDP_PASS
    )

    # --- MỚI: QinQ double VLAN ---
    pkt_qinq = Ether(type=0x88A8) / Dot1Q(vlan=200, type=0x8100) / Dot1Q(vlan=300) / IP() / UDP(sport=10000, dport=80)
    run_xdp_test(
        "TC_1.5", "Bóc tách QinQ (VLAN kép 802.1ad)",
        "Gửi gói bọc 2 lớp VLAN tag (QinQ: outer=200, inner=300).",
        "Bóc 2 lớp VLAN tag, xử lý gói IPv4 UDP bình thường -> PASS",
        raw(pkt_qinq), XDP_PASS
    )

    # --- MỚI: IP header với options (ihl > 5) ---
    # IP header bình thường 20 byte (ihl=5), thêm 4 byte NOP options -> 24 byte (ihl=6)
    pkt_ip_opts = Ether() / IP(ihl=6, options=[b'\x01\x01\x01\x01']) / UDP(sport=10000, dport=80)
    run_xdp_test(
        "TC_1.6", "IP Header với Options (ihl=6)",
        "Gửi gói IP có header dài hơn 20 byte (có IP options).",
        "Xử lý đúng ip_hdr_len = ihl * 4 = 24 -> PASS",
        raw(pkt_ip_opts), XDP_PASS
    )

    # --- MỚI: L4 header truncated (UDP) ---
    # Tạo gói IP+UDP nhưng cắt ngắn để UDP header bị thiếu
    pkt_trunc_udp = raw(Ether() / IP(proto=17) / UDP())[:42]  # Ethernet(14) + IP(20) + chỉ 8 byte -> vừa đủ
    pkt_trunc_udp_short = raw(Ether() / IP(proto=17) / UDP())[:38]  # Cắt bớt 4 byte UDP header
    run_xdp_test(
        "TC_1.7", "UDP Header bị cắt cụt",
        "Gói IP protocol=UDP nhưng thiếu byte tầng L4 (chỉ 4/8 byte UDP hdr).",
        "L4 parse error -> DROP (BUG-1 fix: có track_stats)",
        pkt_trunc_udp_short, XDP_DROP
    )

    # --- MỚI: L4 header truncated (TCP) ---
    pkt_trunc_tcp = raw(Ether() / IP(proto=6) / TCP())[:44]  # TCP header cần 20 byte, chỉ cho 10
    run_xdp_test(
        "TC_1.8", "TCP Header bị cắt cụt",
        "Gói IP protocol=TCP nhưng thiếu byte TCP header (chỉ 10/20 byte).",
        "L4 parse error -> DROP (BUG-1 fix: có track_stats)",
        pkt_trunc_tcp, XDP_DROP
    )

    # --- MỚI: L4 header truncated (ICMP) ---
    pkt_trunc_icmp = raw(Ether() / IP(proto=1) / ICMP())[:38]  # ICMP header cần 8 byte, chỉ cho 4
    run_xdp_test(
        "TC_1.9", "ICMP Header bị cắt cụt",
        "Gói IP protocol=ICMP nhưng thiếu byte ICMP header.",
        "L4 parse error -> DROP (BUG-1 fix: có track_stats)",
        pkt_trunc_icmp, XDP_DROP
    )

    # ===================================================================
    # NHÓM 2: WHITELIST / BLACKLIST
    # ===================================================================
    print("\n" + "=" * 60)
    print("🛡️  NHÓM 2: WHITELIST / BLACKLIST")
    print("=" * 60)

    # IP test dùng cho whitelist/blacklist (dùng IP ít gặp để tránh ảnh hưởng)
    TEST_WL_IP = "198.51.100.1"
    TEST_BL_IP = "198.51.100.2"

    # --- MỚI: Whitelist bypass ---
    print("\n[SETUP] Thêm IP whitelist:", TEST_WL_IP)
    map_add_ip('acl_map', TEST_WL_IP)  # ACL_ALLOW=1 (default value)
    pkt_wl = Ether() / IP(src=TEST_WL_IP) / UDP(sport=53, dport=12345) / (b"A" * 64)
    run_xdp_test(
        "TC_2.1", "Whitelist bypass - UDP Amp port",
        f"Gói UDP từ IP whitelist {TEST_WL_IP}, sport=53 (amp port).",
        "IP whitelisted -> bỏ qua mọi kiểm tra -> PASS (hoặc REDIRECT)",
        raw(pkt_wl), XDP_PASS
    )
    print("[CLEANUP] Xoá IP whitelist:", TEST_WL_IP)
    map_delete_ip('acl_map', TEST_WL_IP)

    # --- MỚI: Blacklist block ---
    print("\n[SETUP] Thêm IP blacklist:", TEST_BL_IP)
    map_add_ip('acl_map', TEST_BL_IP)  # Note: need to set value=0x02 for ACL_DENY
    pkt_bl = Ether() / IP(src=TEST_BL_IP) / TCP(sport=12345, dport=80, flags="A")
    run_xdp_test(
        "TC_2.3", "Blacklist chặn TCP bình thường",
        f"Gói TCP ACK hợp lệ từ IP blacklisted {TEST_BL_IP}.",
        "IP blacklisted -> DROP ngay, ưu tiên trước whitelist",
        raw(pkt_bl), XDP_DROP
    )
    print("[CLEANUP] Xoá IP blacklist:", TEST_BL_IP)
    map_delete_ip('acl_map', TEST_BL_IP)

    # ===================================================================
    # NHÓM 3: IP FRAGMENT
    # ===================================================================
    print("\n" + "=" * 60)
    print("💥 NHÓM 3: IP FRAGMENT")
    print("=" * 60)

    pkt_frag = Ether() / IP(flags="MF", frag=0) / UDP(sport=1234, dport=80)
    run_xdp_test(
        "TC_3.1", "Chặn IP Fragment (MF flag)",
        "Gói IP có cờ More Fragments (MF=1).",
        "Phát hiện phân mảnh -> DROP",
        raw(pkt_frag), XDP_DROP
    )

    # Fragment offset > 0 (gói tin phân mảnh phần tiếp theo)
    pkt_frag_offset = Ether() / IP(frag=100) / (b"\x00" * 20)
    run_xdp_test(
        "TC_3.2", "Chặn IP Fragment (offset > 0)",
        "Gói IP có fragment offset = 100 (fragment tiếp theo).",
        "Phát hiện frag_off có IP_OFFSET -> DROP",
        raw(pkt_frag_offset), XDP_DROP
    )

    # ===================================================================
    # NHÓM 4: UDP CHECKS
    # ===================================================================
    print("\n" + "=" * 60)
    print("📡 NHÓM 4: UDP CHECKS")
    print("=" * 60)

    # Setup: đảm bảo port 53 nằm trong amp_ports_map
    print("\n[SETUP] Thêm port 53 vào amp_ports_map")
    map_add_port('amp_ports_map', 53)

    pkt_dns = Ether() / IP() / UDP(sport=53, dport=12345) / (b"A" * 64)
    run_xdp_test(
        "TC_4.1", "Chặn UDP Amplification (Port 53)",
        "Gói UDP từ port 53 (DNS).",
        "Lookup amp_ports_map thấy port bị chặn -> DROP",
        raw(pkt_dns), XDP_DROP
    )
    
    # --- MỚI: UDP từ port KHÔNG trong amp_ports_map -> PASS ---
    pkt_udp_normal = Ether() / IP() / UDP(sport=10000, dport=80) / (b"B" * 100)
    run_xdp_test(
        "TC_4.2", "UDP từ port hợp lệ (non-amp)",
        "Gói UDP từ port 10000, payload 100 bytes.",
        "Port không bị chặn, payload < 1024 -> PASS",
        raw(pkt_udp_normal), XDP_PASS
    )

    pkt_huge_udp = Ether() / IP() / UDP(sport=12345, dport=80) / (b"X" * 1500)
    run_xdp_test(
        "TC_4.3", "Chặn UDP Payload > 1024",
        "Gói UDP payload 1500 bytes.",
        "Vượt quá config (default 1024) -> DROP",
        raw(pkt_huge_udp), XDP_DROP
    )

    # --- MỚI: Boundary test - payload ĐÚNG BẰNG max_size (1024) -> PASS ---
    pkt_boundary_ok = Ether() / IP() / UDP(sport=12345, dport=80) / (b"Y" * 1024)
    run_xdp_test(
        "TC_4.4", "UDP Payload = 1024 (boundary PASS)",
        "Gói UDP payload đúng 1024 bytes (= config max_size).",
        "payload_len == max_size -> KHÔNG vượt ngưỡng -> PASS",
        raw(pkt_boundary_ok), XDP_PASS
    )

    # --- MỚI: Boundary test - payload = max_size + 1 (1025) -> DROP ---
    pkt_boundary_drop = Ether() / IP() / UDP(sport=12345, dport=80) / (b"Z" * 1025)
    run_xdp_test(
        "TC_4.5", "UDP Payload = 1025 (boundary DROP)",
        "Gói UDP payload 1025 bytes (= config max_size + 1).",
        "payload_len > max_size -> DROP",
        raw(pkt_boundary_drop), XDP_DROP
    )

    # ===================================================================
    # NHÓM 5: TCP CHECKS
    # ===================================================================
    print("\n" + "=" * 60)
    print("🔗 NHÓM 5: TCP CHECKS")
    print("=" * 60)

    pkt_syn = Ether() / IP() / TCP(sport=12345, dport=80, flags="S")
    run_xdp_test(
        "TC_5.1", "Gói TCP SYN hợp lệ",
        "Gửi 1 gói SYN thuần.",
        "Dưới rate limit -> PASS",
        raw(pkt_syn), XDP_PASS
    )

    # --- MỚI: TCP SYN+FIN (invalid combination) ---
    pkt_synfin = Ether() / IP() / TCP(sport=12345, dport=80, flags="SF")
    run_xdp_test(
        "TC_5.2", "TCP SYN+FIN (flag bất thường)",
        "Gói TCP có cả SYN và FIN set cùng lúc.",
        "Tổ hợp flag không hợp lệ (scan/attack) -> DROP",
        raw(pkt_synfin), XDP_DROP
    )

    # --- MỚI: TCP SYN+RST (invalid combination) ---
    pkt_synrst = Ether() / IP() / TCP(sport=12345, dport=80, flags="SR")
    run_xdp_test(
        "TC_5.3", "TCP SYN+RST (flag bất thường)",
        "Gói TCP có cả SYN và RST set cùng lúc.",
        "Tổ hợp flag không hợp lệ -> DROP",
        raw(pkt_synrst), XDP_DROP
    )

    # --- MỚI: TCP NULL scan (flags = 0) ---
    pkt_null = Ether() / IP() / TCP(sport=12345, dport=80, flags=0)
    run_xdp_test(
        "TC_5.4", "TCP NULL Scan (flags = 0)",
        "Gói TCP không set bất kỳ flag nào (null scan/Nmap).",
        "flags == 0 -> phát hiện scan -> DROP",
        raw(pkt_null), XDP_DROP
    )

    # --- MỚI: TCP ACK thuần (established connection) ---
    pkt_ack = Ether() / IP() / TCP(sport=12345, dport=80, flags="A")
    run_xdp_test(
        "TC_5.5", "TCP ACK hợp lệ (established)",
        "Gói TCP ACK thuần (traffic kết nối đã thiết lập).",
        "Flag hợp lệ, không phải SYN -> PASS (không rate limit)",
        raw(pkt_ack), XDP_PASS
    )

    # --- MỚI: TCP SYN-ACK (bước 2 của 3-way handshake) ---
    pkt_synack = Ether() / IP() / TCP(sport=80, dport=12345, flags="SA")
    run_xdp_test(
        "TC_5.6", "TCP SYN-ACK (handshake bước 2)",
        "Gói TCP SYN-ACK (response từ server).",
        "SYN+ACK -> không bị rate limit (chỉ SYN thuần mới bị) -> PASS",
        raw(pkt_synack), XDP_PASS
    )

    # --- MỚI: TCP FIN-ACK (đóng kết nối) ---
    pkt_finack = Ether() / IP() / TCP(sport=12345, dport=80, flags="FA")
    run_xdp_test(
        "TC_5.7", "TCP FIN-ACK (đóng kết nối)",
        "Gói TCP FIN-ACK (đóng kết nối bình thường).",
        "Flag hợp lệ -> PASS",
        raw(pkt_finack), XDP_PASS
    )

    # --- MỚI: TCP RST (reset kết nối) ---
    pkt_rst = Ether() / IP() / TCP(sport=12345, dport=80, flags="R")
    run_xdp_test(
        "TC_5.8", "TCP RST (reset kết nối)",
        "Gói TCP RST đơn lẻ.",
        "RST flag hợp lệ, không bị phát hiện bất thường -> PASS",
        raw(pkt_rst), XDP_PASS
    )

    # ===================================================================
    # NHÓM 6: ICMP CHECKS
    # ===================================================================
    print("\n" + "=" * 60)
    print("🏓 NHÓM 6: ICMP CHECKS")
    print("=" * 60)

    pkt_icmp = Ether() / IP() / ICMP(type=8)
    run_xdp_test(
        "TC_6.1", "Gói ICMP Echo Request (Ping)",
        "Gửi 1 gói Ping (ICMP type 8).",
        "Dưới rate limit -> PASS",
        raw(pkt_icmp), XDP_PASS
    )

    # --- MỚI: ICMP Echo Reply ---
    pkt_icmp_reply = Ether() / IP() / ICMP(type=0)
    run_xdp_test(
        "TC_6.2", "Gói ICMP Echo Reply",
        "Gửi 1 gói ICMP Echo Reply (type 0).",
        "Dưới rate limit -> PASS",
        raw(pkt_icmp_reply), XDP_PASS
    )

    # ===================================================================
    # NHÓM 7: GIAO THỨC KHÁC
    # ===================================================================
    print("\n" + "=" * 60)
    print("🌐 NHÓM 7: GIAO THỨC KHÁC")
    print("=" * 60)

    # --- MỚI: GRE (protocol 47) ---
    # Craft gói IP với protocol=47 (GRE), payload giả
    pkt_gre = Ether() / IP(proto=47) / (b"\x00" * 20)
    run_xdp_test(
        "TC_7.1", "Giao thức GRE (proto=47)",
        "Gói IP với protocol GRE tunnel.",
        "Không có rule đặc biệt cho GRE -> PASS (qua kernel xử lý)",
        raw(pkt_gre), XDP_PASS
    )

    # --- MỚI: ESP (protocol 50 - IPSec) ---
    pkt_esp = Ether() / IP(proto=50) / (b"\x00" * 20)
    run_xdp_test(
        "TC_7.2", "Giao thức ESP/IPSec (proto=50)",
        "Gói IP với protocol ESP (VPN tunnel).",
        "Không có rule đặc biệt cho ESP -> PASS",
        raw(pkt_esp), XDP_PASS
    )

    # ===================================================================
    # TỔNG KẾT
    # ===================================================================
    print("\n" + "=" * 60)
    total = len(test_results)
    passed = sum(1 for r in test_results if r["Trạng Thái"] == "PASS")
    failed = sum(1 for r in test_results if r["Trạng Thái"] == "FAIL")
    errors = sum(1 for r in test_results if r["Trạng Thái"] == "ERROR")
    print(f"📊 TỔNG KẾT: {total} tests | ✔ PASS: {passed} | ✘ FAIL: {failed} | ⚠ ERROR: {errors}")
    print("=" * 60)


if __name__ == "__main__":
    run_all_tests()
    export_excel_report()